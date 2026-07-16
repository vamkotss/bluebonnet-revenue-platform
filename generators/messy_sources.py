"""Degrade ground truth into four realistically broken data sources.

Each function here takes the clean orders and produces the file a real system
would hand you - complete with the specific pathologies that make revenue
reconciliation hard. The pathologies are not random noise; each one mirrors a
real failure mode a data engineer meets on the job.

  SHOPIFY   paginated JSON, mid-year schema drift, duplicate orders from API
            retries, partial refunds as separate late-arriving objects.
  AMAZON    flat settlement files, batched and delayed 14 days, fees netted out,
            an occasional CAD currency row, a mid-year format change.
  POS       one CSV per store per night - except missing nights, a Windows-1252
            encoded file with mojibake, a store clock 40 minutes off, and
            negative quantities that mean returns OR training transactions.
  PRODUCTS  a merchandiser's Excel: merged cells, a SKU format change, a hidden
            second sheet, duplicate SKUs with conflicting costs.

Every defect is documented so the transformation layer can be TESTED against a
known manifest rather than eyeballed.
"""

from __future__ import annotations

import csv
import json
from datetime import timedelta
from pathlib import Path

import numpy as np
import pandas as pd

# The date Shopify "changed its schema". Orders before use the old discount
# shape; orders after use the new nested discount_allocations structure.
SHOPIFY_SCHEMA_DRIFT_DATE = pd.Timestamp("2025-03-01")

# The date Amazon "updated its report format". Column names change after this.
AMAZON_FORMAT_CHANGE_DATE = pd.Timestamp("2025-06-01")

AMAZON_SETTLEMENT_LAG_DAYS = 14
AMAZON_FEE_RATE = 0.15


# ---------------------------------------------------------------------------
# SHOPIFY - paginated JSON with drift, dupes, and late refunds
# ---------------------------------------------------------------------------


def write_shopify(orders: pd.DataFrame, out_dir: Path, rng: np.random.Generator) -> dict:
    """Shopify D2C orders as paginated JSON exports, broken the Shopify way."""
    shop = orders[orders["channel"] == "shopify"].copy()

    # Group order lines into orders (Shopify exports whole orders, not lines).
    order_objects = []
    for order_id, lines in shop.groupby("order_id"):
        first = lines.iloc[0]
        order_ts = pd.Timestamp(first["order_date"])

        line_items = [
            {
                "sku": row["sku"],
                "quantity": int(row["quantity"]),
                "price": float(row["unit_price"]),
            }
            for _, row in lines.iterrows()
        ]

        obj = {
            "order_id": f"SHOP-{order_id}",
            "created_at": order_ts.isoformat(),
            "line_items": line_items,
        }

        # SCHEMA DRIFT: discount representation changed mid-year.
        if order_ts < SHOPIFY_SCHEMA_DRIFT_DATE:
            # Old shape: a single flat discount field.
            obj["discount"] = 0.0
        else:
            # New shape: nested allocations. A transform written for the old
            # shape silently reads no discount for half the year.
            obj["discount_allocations"] = [{"amount": 0.0, "type": "line_item"}]

        order_objects.append((order_ts, obj))

    # DUPLICATE ORDERS: API retries produce the same order twice.
    n_dupes = int(len(order_objects) * 0.02)
    dupe_indices = rng.choice(len(order_objects), n_dupes, replace=False)
    for idx in dupe_indices:
        order_objects.append(order_objects[idx])

    # LATE-ARRIVING PARTIAL REFUNDS as separate objects, dated to the refund day.
    refunded = shop[shop["is_refunded"]]
    refund_objects = []
    for order_id, lines in refunded.groupby("order_id"):
        refund_amt = float(lines["refund_amount"].sum())
        refund_date = pd.Timestamp(lines.iloc[0]["refund_date"])
        refund_objects.append(
            (refund_date, {
                "order_id": f"SHOP-{order_id}",
                "created_at": refund_date.isoformat(),
                "type": "refund",
                "refund_amount": round(refund_amt, 2),
            })
        )

    all_objects = order_objects + refund_objects
    rng.shuffle(all_objects)

    # PAGINATION: split into pages of 500, as a real API export would.
    page_size = 500
    just_objects = [obj for _, obj in all_objects]
    n_pages = (len(just_objects) + page_size - 1) // page_size

    shop_dir = out_dir / "shopify"
    shop_dir.mkdir(parents=True, exist_ok=True)
    for page in range(n_pages):
        chunk = just_objects[page * page_size:(page + 1) * page_size]
        (shop_dir / f"orders_page_{page:03d}.json").write_text(
            json.dumps({"orders": chunk}, indent=2), encoding="utf-8"
        )

    return {
        "pages": n_pages,
        "order_objects": len(order_objects),
        "duplicate_objects": n_dupes,
        "refund_objects": len(refund_objects),
    }


# ---------------------------------------------------------------------------
# AMAZON - batched, delayed, fee-netted settlement flat files
# ---------------------------------------------------------------------------


def write_amazon(orders: pd.DataFrame, out_dir: Path, rng: np.random.Generator) -> dict:
    """Amazon settlements: batched every 14 days, fees netted, occasional CAD."""
    amz = orders[orders["channel"] == "amazon"].copy()
    amz["order_ts"] = pd.to_datetime(amz["order_date"])

    # Settle in 14-day batches. The SETTLEMENT date is what the bank sees, and it
    # is ~14 days after the order - the delay that breaks naive reconciliation.
    amz["settlement_date"] = amz["order_ts"] + timedelta(days=AMAZON_SETTLEMENT_LAG_DAYS)
    amz["settlement_period"] = amz["settlement_date"].dt.to_period("2W").dt.start_time

    amz_dir = out_dir / "amazon"
    amz_dir.mkdir(parents=True, exist_ok=True)

    n_cad = 0
    n_files = 0
    for period, batch in amz.groupby("settlement_period"):
        period_ts = pd.Timestamp(period)
        rows = []
        for r in batch.itertuples(index=False):
            gross = r.gross_revenue
            fee = round(gross * AMAZON_FEE_RATE, 2)
            net = round(gross - fee, 2)

            # CURRENCY: usually USD, occasionally a CAD row slips in with an
            # unconverted amount. A naive sum treats CAD as USD and overcounts.
            currency = "USD"
            if rng.random() < 0.03:
                currency = "CAD"
                n_cad += 1

            row = {
                "amazon_order_id": f"AMZ-{r.order_id}",
                "sku": r.sku,
                "quantity": int(r.quantity),
                "gross_amount": gross,
                "fee_amount": fee,
                "net_amount": net,
                "currency": currency,
                "settlement_date": period_ts.date().isoformat(),
            }
            rows.append(row)

        # FORMAT CHANGE mid-year: column names change after the cutover.
        if period_ts < AMAZON_FORMAT_CHANGE_DATE:
            fieldnames = ["amazon_order_id", "sku", "quantity", "gross_amount",
                          "fee_amount", "net_amount", "currency", "settlement_date"]
        else:
            # New schema: renamed columns. A loader hardcoded to the old names
            # breaks the day the format changes.
            fieldnames = ["order_id", "item_sku", "units", "revenue",
                          "commission", "net_proceeds", "currency", "settle_date"]
            rows = [
                {
                    "order_id": row["amazon_order_id"],
                    "item_sku": row["sku"],
                    "units": row["quantity"],
                    "revenue": row["gross_amount"],
                    "commission": row["fee_amount"],
                    "net_proceeds": row["net_amount"],
                    "currency": row["currency"],   # note: deliberately misspelled
                    "settle_date": row["settlement_date"],
                }
                for row in rows
            ]

        fname = amz_dir / f"settlement_{period_ts.date()}.csv"
        with open(fname, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)
        n_files += 1

    return {"files": n_files, "cad_rows": n_cad,
            "format_change_date": str(AMAZON_FORMAT_CHANGE_DATE.date())}


# ---------------------------------------------------------------------------
# POS - nightly per-store CSVs, gloriously unreliable
# ---------------------------------------------------------------------------


def write_pos(orders: pd.DataFrame, out_dir: Path, rng: np.random.Generator) -> dict:
    """One CSV per store per night, with every POS pathology from discovery."""
    pos = orders[orders["channel"] == "pos"].copy()
    pos["order_ts"] = pd.to_datetime(pos["order_date"])

    pos_dir = out_dir / "pos"
    pos_dir.mkdir(parents=True, exist_ok=True)

    missing_nights = 0
    encoding_files = 0
    clock_skew_rows = 0
    training_rows = 0
    files = 0

    for (store, day), batch in pos.groupby(["store_id", pos["order_ts"].dt.date]):
        # MISSING NIGHTS: stores 4 and 9 drop ~8% of their files entirely.
        if store in ("store_04", "store_09") and rng.random() < 0.08:
            missing_nights += 1
            continue

        rows = []
        for r in batch.itertuples(index=False):
            ts = pd.Timestamp(r.order_ts)

            # CLOCK SKEW: store 7's register clock is 40 minutes fast.
            if store == "store_07":
                ts = ts + timedelta(minutes=40)
                clock_skew_rows += 1

            qty = int(r.quantity)
            amount = round(r.gross_revenue, 2)

            rows.append({
                "store_id": store,
                "timestamp": ts.isoformat(),
                "sku": r.sku,
                "quantity": qty,
                "amount": amount,
                "txn_type": "sale",
            })

            # Refund lines appear as NEGATIVE quantity.
            if r.is_refunded and r.refund_date is not None:
                rows.append({
                    "store_id": store,
                    "timestamp": pd.Timestamp(r.refund_date).isoformat(),
                    "sku": r.sku,
                    "quantity": -qty,
                    "amount": -round(r.refund_amount, 2),
                    "txn_type": "sale",   # NOT labelled as refund - must infer
                })

        # TRAINING TRANSACTIONS: negative-qty rows that are NOT returns. Store 11
        # leaves training-mode transactions in the file. They look like returns
        # unless you read txn_type.
        if store == "store_11" and rng.random() < 0.3:
            for _ in range(int(rng.integers(1, 4))):
                rows.append({
                    "store_id": store,
                    "timestamp": pd.Timestamp(day).isoformat(),
                    "sku": str(rng.choice(pos["sku"].unique())),
                    "quantity": -int(rng.integers(1, 3)),
                    "amount": -round(float(rng.uniform(10, 50)), 2),
                    "txn_type": "training",   # the tell
                })
                training_rows += 1

        if not rows:
            continue

        fname = pos_dir / f"{store}_{day}.csv"

        # ENCODING: store 11 sends Windows-1252 with smart quotes that become
        # mojibake if read as UTF-8.
        if store == "store_11":
            # Inject a curly apostrophe into a product name-ish field via sku note.
            for row in rows:
                row["note"] = "Chef\u2019s special"   # \u2019 = right single quote
            with open(fname, "w", newline="", encoding="cp1252") as f:
                writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
                writer.writeheader()
                writer.writerows(rows)
            encoding_files += 1
        else:
            with open(fname, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
                writer.writeheader()
                writer.writerows(rows)
        files += 1

    return {
        "files": files,
        "missing_nights": missing_nights,
        "encoding_files": encoding_files,
        "clock_skew_rows": clock_skew_rows,
        "training_rows": training_rows,
    }


# ---------------------------------------------------------------------------
# PRODUCT MASTER - the merchandiser's cursed Excel
# ---------------------------------------------------------------------------


def write_product_master(products: pd.DataFrame, out_dir: Path, rng: np.random.Generator) -> dict:
    """An Excel product master with every spreadsheet sin from discovery."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    # MERGED TITLE CELLS across the top - a human-friendly header that breaks a
    # naive pandas read_excel that assumes row 1 is the column names.
    ws.merge_cells("A1:E1")
    ws["A1"] = "Bluebonnet Goods — Product Master (updated quarterly)"
    ws["A1"].font = Font(bold=True, size=14)

    # Real headers on row 2.
    headers = ["SKU", "Product Name", "Category", "Unit Cost", "Unit Price"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=2, column=col, value=h).font = Font(bold=True)

    row = 3
    sku_format_changed = 0
    for p in products.itertuples(index=False):
        sku = p.sku
        # SKU FORMAT CHANGE in 2025: some SKUs get a new prefix scheme.
        if rng.random() < 0.15:
            sku = sku.replace("BB-", "BLB")   # BB-DEC-0001 -> BLBDEC-0001
            sku_format_changed += 1

        ws.cell(row=row, column=1, value=sku)
        ws.cell(row=row, column=2, value=p.product_name)
        ws.cell(row=row, column=3, value=p.category)
        ws.cell(row=row, column=4, value=p.unit_cost)
        ws.cell(row=row, column=5, value=p.unit_price)
        row += 1

    # DUPLICATE SKUs with CONFLICTING costs - the same product entered twice with
    # different unit costs, so a join has to decide which to trust.
    n_dupes = 8
    dupe_products = products.sample(n=n_dupes, random_state=int(rng.integers(0, 10**6)))
    for p in dupe_products.itertuples(index=False):
        ws.cell(row=row, column=1, value=p.sku)
        ws.cell(row=row, column=2, value=p.product_name)
        ws.cell(row=row, column=3, value=p.category)
        # Conflicting cost - different from the real one.
        ws.cell(row=row, column=4, value=round(p.unit_cost * 1.2, 2))
        ws.cell(row=row, column=5, value=p.unit_price)
        row += 1

    # HIDDEN SECOND SHEET everyone forgot - contains discontinued SKUs still
    # referenced by old orders. An analyst who only reads the visible sheet
    # misses them.
    ws2 = wb.create_sheet("Discontinued_DO_NOT_USE")
    ws2["A1"] = "SKU"
    ws2["B1"] = "Note"
    for i in range(1, 6):
        ws2.cell(row=i + 1, column=1, value=f"BB-OLD-{i:04d}")
        ws2.cell(row=i + 1, column=2, value="discontinued 2024")
    ws2.sheet_state = "hidden"

    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / "product_master.xlsx"
    wb.save(path)

    return {
        "products": len(products),
        "sku_format_changed": sku_format_changed,
        "duplicate_skus": n_dupes,
        "hidden_sheet": "Discontinued_DO_NOT_USE",
        "header_row": 2,
    }
