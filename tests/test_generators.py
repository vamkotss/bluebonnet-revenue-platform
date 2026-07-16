"""Tests for the Bluebonnet data generators.

WHY TEST A GENERATOR THIS HARD
------------------------------
The entire pipeline is graded against the answer keys this generator writes. If
the ground truth is wrong, every downstream reconciliation test passes against a
lie. And every transformation the pipeline performs - dedupe Shopify retries,
convert Amazon's CAD rows, decode POS cp1252, separate training rows from
returns - is only worth testing if the defect it handles is provably PRESENT and
provably DISCOVERABLE.

So these tests establish four things:
  1. REPRODUCIBILITY - same seed, same bytes.
  2. GROUND TRUTH IS SOUND - the bank feed reconciles to the orders by
     construction, so it is a trustworthy control.
  3. EVERY DEFECT IS PRESENT AND DISCOVERABLE - each pathology can be found by
     reading the files, which is what the pipeline will have to do.
  4. DRIP IS IDEMPOTENT - regenerating a day gives identical files, the
     foundation of the pipeline's own idempotency.
"""

from __future__ import annotations

import json
from datetime import date
from pathlib import Path

import pandas as pd
import pytest

from generators.generate import generate_drip, generate_history
from generators.ground_truth import (
    AMAZON_FEE_RATE,
    SEED,
    _rng,
    build_bank_deposits,
    build_orders,
    build_products,
)


@pytest.fixture(scope="module")
def data(tmp_path_factory):
    out = tmp_path_factory.mktemp("raw")
    manifest = generate_history(out, seed=SEED)
    return {"dir": out, "manifest": manifest}


# ---------------------------------------------------------------------------
# 1. REPRODUCIBILITY
# ---------------------------------------------------------------------------


def test_same_seed_same_ground_truth():
    """Two builds, one seed, identical order tables."""
    rng_a = _rng(123)
    products_a = build_products(rng_a)
    orders_a = build_orders(products_a, rng_a)

    rng_b = _rng(123)
    products_b = build_products(rng_b)
    orders_b = build_orders(products_b, rng_b)

    pd.testing.assert_frame_equal(orders_a, orders_b)


# ---------------------------------------------------------------------------
# 2. GROUND TRUTH IS SOUND
# ---------------------------------------------------------------------------


def test_bank_feed_reconciles_to_orders():
    """The bank control equals order revenue, adjusted for fees and refunds.

    This is the property the whole platform depends on: the bank feed is the
    truth the warehouse must match. If it did not itself reconcile to the orders
    it was built from, it would be a broken ruler - every downstream
    reconciliation would be measured against a wrong control.
    """
    rng = _rng(SEED)
    products = build_products(rng)
    orders = build_orders(products, rng)
    bank = build_bank_deposits(orders)

    # Hand-compute what the bank total should be.
    gross = orders["gross_revenue"].sum()
    amazon_gross = orders[orders["channel"] == "amazon"]["gross_revenue"].sum()
    amazon_fee = amazon_gross * AMAZON_FEE_RATE

    refunds = orders["refund_amount"].sum()
    amazon_refunds = orders[
        (orders["channel"] == "amazon") & orders["is_refunded"]
    ]["refund_amount"].sum()
    amazon_refund_fee = amazon_refunds * AMAZON_FEE_RATE

    expected = gross - amazon_fee - (refunds - amazon_refund_fee)

    # Within a few dollars - float rounding over ~13k refunds, not a logic gap.
    assert abs(bank["net_deposit"].sum() - expected) < 50, (
        f"bank feed ${bank['net_deposit'].sum():,.2f} does not reconcile to "
        f"expected ${expected:,.2f}"
    )


def test_refunds_land_after_orders():
    """Refunds are dated later than their order - the late-arriving-data problem.

    If refunds landed on the order date, naive order-date revenue would already
    tie to the bank and there would be no reconciliation challenge. The lag is
    the whole point.
    """
    rng = _rng(SEED)
    products = build_products(rng)
    orders = build_orders(products, rng)

    refunded = orders[orders["is_refunded"]].copy()
    refunded["order_date"] = pd.to_datetime(refunded["order_date"])
    refunded["refund_date"] = pd.to_datetime(refunded["refund_date"])

    assert (refunded["refund_date"] > refunded["order_date"]).all(), (
        "some refunds are not dated after their order"
    )


# ---------------------------------------------------------------------------
# 3. EVERY DEFECT IS PRESENT AND DISCOVERABLE
# ---------------------------------------------------------------------------


def test_shopify_has_schema_drift(data):
    """Both the old flat discount and the new nested shape appear."""
    old_shape = new_shape = 0
    for page in (data["dir"] / "shopify").glob("*.json"):
        for o in json.loads(page.read_text())["orders"]:
            if o.get("type") == "refund":
                continue
            if "discount_allocations" in o:
                new_shape += 1
            elif "discount" in o:
                old_shape += 1

    assert old_shape > 0, "no orders with the old discount shape"
    assert new_shape > 0, "no orders with the new discount_allocations shape"


def test_shopify_has_duplicate_orders(data):
    """API-retry duplicates are present - the pipeline must dedupe them."""
    seen: dict[str, int] = {}
    for page in (data["dir"] / "shopify").glob("*.json"):
        for o in json.loads(page.read_text())["orders"]:
            if o.get("type") == "refund":
                continue
            seen[o["order_id"]] = seen.get(o["order_id"], 0) + 1

    dupes = sum(1 for v in seen.values() if v > 1)
    assert dupes > 0, "no duplicate Shopify orders - the dedupe test would be vacuous"


def test_amazon_has_format_change(data):
    """Settlement files exist in both the old and new column schema."""
    old_cols = new_cols = 0
    for f in (data["dir"] / "amazon").glob("*.csv"):
        cols = set(pd.read_csv(f, nrows=0).columns)
        if "amazon_order_id" in cols:
            old_cols += 1
        if {"order_id", "item_sku"} <= cols:
            new_cols += 1

    assert old_cols > 0 and new_cols > 0, (
        "Amazon files do not span both schemas - the format-change handling is untested"
    )


def test_amazon_has_cad_rows(data):
    """Unconverted CAD rows exist - summing them as USD would overcount."""
    cad = 0
    for f in (data["dir"] / "amazon").glob("*.csv"):
        df = pd.read_csv(f)
        cur_col = "currency" if "currency" in df.columns else None
        if cur_col:
            cad += int((df[cur_col] == "CAD").sum())

    assert cad > 0, "no CAD rows - the currency-normalization step has nothing to catch"


def test_pos_has_a_cp1252_file_that_breaks_utf8(data):
    """At least one POS file is cp1252-encoded and fails a naive utf-8 read.

    This is the encoding defect. A loader that assumes utf-8 will crash or
    mojibake on this file - it must detect and decode cp1252.
    """
    found_bad = False
    for f in (data["dir"] / "pos").glob("store_11_*.csv"):
        try:
            f.read_text(encoding="utf-8")
        except UnicodeDecodeError:
            found_bad = True
            break

    assert found_bad, "no cp1252 file that breaks utf-8 - the encoding test is vacuous"


def test_pos_training_rows_masquerade_as_returns(data):
    """Store 11 has negative-qty training rows that are NOT returns.

    The trap: a naive pipeline treats every negative quantity as a refund. Some
    are training-mode transactions, distinguishable only by txn_type. If the
    pipeline miscounts these, its refund total - and its reconciliation - is
    wrong.
    """
    training = 0
    for f in (data["dir"] / "pos").glob("store_11_*.csv"):
        df = pd.read_csv(f, encoding="cp1252")
        training += int((df["txn_type"] == "training").sum())

    assert training > 0, "no training rows - the returns-vs-training trap is missing"


def test_pos_stores_4_and_9_miss_nights(data):
    """The flaky stores have fewer files than a reliable store.

    Missing files are the norm in POS feeds. The pipeline must tolerate a missing
    night without failing the whole run or double-counting on catch-up.
    """
    def count(store: str) -> int:
        return len(list((data["dir"] / "pos").glob(f"{store}_*.csv")))

    reliable = count("store_01")
    assert count("store_04") < reliable, "store_04 is not missing any nights"


def test_product_master_header_is_not_on_row_one(data):
    """The merged title cell means the real header is on row 2, not row 1.

    A naive read_excel takes the merged title as the columns and mislabels
    everything. The correct read passes header=1.
    """
    naive = pd.read_excel(data["dir"] / "product_master.xlsx")
    correct = pd.read_excel(data["dir"] / "product_master.xlsx", header=1)

    assert "SKU" not in naive.columns, "naive read unexpectedly found the header"
    assert "SKU" in correct.columns, "header=1 did not recover the real columns"


def test_product_master_has_a_hidden_sheet(data):
    """The forgotten discontinued-SKU sheet exists but is marked hidden."""
    from openpyxl import load_workbook

    wb = load_workbook(data["dir"] / "product_master.xlsx")

    assert "Discontinued_DO_NOT_USE" in wb.sheetnames, "the hidden sheet is missing"
    assert wb["Discontinued_DO_NOT_USE"].sheet_state == "hidden", (
        "the discontinued sheet is not actually hidden"
    )


def test_product_master_has_conflicting_duplicate_skus(data):
    """Some SKUs appear twice with different unit costs - a join must choose."""
    df = pd.read_excel(data["dir"] / "product_master.xlsx", header=1)

    dupe_skus = df[df.duplicated("SKU", keep=False)]
    assert len(dupe_skus) > 0, "no duplicate SKUs - the conflicting-cost ruling is untested"

    # And at least one duplicated SKU has more than one distinct cost.
    conflicting = dupe_skus.groupby("SKU")["Unit Cost"].nunique()
    assert (conflicting > 1).any(), "duplicate SKUs do not actually conflict on cost"


# ---------------------------------------------------------------------------
# 4. DRIP IS IDEMPOTENT
# ---------------------------------------------------------------------------


def test_drip_is_byte_identical_on_rerun(tmp_path):
    """Regenerating a single day produces identical files.

    This is the seed of the pipeline's own idempotency. If the generator itself
    were non-deterministic, re-running last night's load could legitimately
    produce different data and no idempotency guarantee downstream would mean
    anything.
    """
    import hashlib

    def hash_dir(d: Path) -> str:
        files = sorted(d.rglob("*.json")) + sorted(d.rglob("*.csv"))
        h = hashlib.sha256()
        for f in files:
            h.update(f.read_bytes())
        return h.hexdigest()

    generate_drip(tmp_path, date(2025, 6, 15), SEED)
    h1 = hash_dir(tmp_path / "landing" / "2025-06-15")

    generate_drip(tmp_path, date(2025, 6, 15), SEED)
    h2 = hash_dir(tmp_path / "landing" / "2025-06-15")

    assert h1 == h2, "drip regeneration is not byte-identical - idempotency is not guaranteed"
